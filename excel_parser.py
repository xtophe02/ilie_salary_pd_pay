from __future__ import annotations

"""
Excel parser for monthly payroll files.

Supports two layouts observed in the reference files:

Layout A (February 2026 — the canonical format):
  Row 1: blank
  Row 2: headers  →  SEN | SALARY <Month> <Year> | PERDIEM <Month> <Year> | LEGAL BONUS … | COMMENTS | …
  Row 3+: data

Layout B (December 2025 — older format with IBAN in col-C instead of perdiem):
  Row 1: headers  →  SEN | SALARY … | <IBAN col> | COMMENTS | <VLOOKUP IBAN>
  Row 2+: data

The parser auto-detects the layout by inspecting column headers and sampled cell values.
"""

import re
from typing import Optional
import openpyxl

MONTH_NAMES = {
    'JANUARY': 1, 'FEBRUARY': 2, 'MARCH': 3, 'APRIL': 4,
    'MAY': 5, 'JUNE': 6, 'JULY': 7, 'AUGUST': 8,
    'SEPTEMBER': 9, 'OCTOBER': 10, 'NOVEMBER': 11, 'DECEMBER': 12,
}

# Minimal IBAN sanity check: 2 letters, 2 digits, then alphanumeric, total 15-34 chars
_IBAN_RE = re.compile(r'^[A-Z]{2}\d{2}[A-Z0-9]{11,30}$')


def _clean_iban(value) -> Optional[str]:
    if not value:
        return None
    cleaned = re.sub(r'\s+', '', str(value)).upper()
    if _IBAN_RE.match(cleaned):
        return cleaned
    return None


def _is_iban(value) -> bool:
    return _clean_iban(value) is not None


def _is_positive_number(value) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool) and value > 0


def _detect_month_year(headers: list) -> tuple[int, int]:
    """Extract month and year from column header strings like 'SALARY FEBRUARY 2026'."""
    for h in headers:
        if not h:
            continue
        h_up = str(h).upper()
        for month_name, month_num in MONTH_NAMES.items():
            if month_name in h_up:
                m = re.search(r'(\d{4})', h_up)
                year = int(m.group(1)) if m else 2026
                return month_num, year
    return 1, 2026  # fallback


def _col_header_type(header_str: str) -> str:
    """Classify a column header into: name | salary | perdiem | bonus | iban | unknown."""
    h = header_str.upper() if header_str else ''
    if any(k in h for k in ('SEN', 'NAME', 'EMPLOYEE', 'SALARIAT')):
        return 'name'
    if 'SALARY' in h or 'SALARIU' in h or 'WAGE' in h:
        return 'salary'
    if any(k in h for k in ('PERDIEM', 'PER DIEM', 'ALLOWANCE', 'DIURNA', 'INDEMNIZATIE')):
        return 'perdiem'
    if any(k in h for k in ('BONUS', 'LEGAL', 'PRIMA')):
        return 'bonus'
    if any(k in h for k in ('IBAN', 'BANK', 'ACCOUNT', 'CONT', 'COMMENT')):
        return 'iban'
    return 'unknown'


def parse_excel(filepath: str) -> dict:
    """
    Parse an Excel payroll file.

    Returns:
        {
          'month': int,
          'year': int,
          'salary_records': [
              {'name': str, 'amount': float, 'iban': str|None, 'index': int}
          ],
          'perdiem_records': [
              {'name': str, 'amount': float, 'iban': str|None, 'salary_index': int}
          ],
          'bonus_records': [
              {'name': str, 'amount': float, 'iban': str|None, 'salary_index': int}
          ],
        }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # ── 1. Find header row (first row that contains recognisable payroll keywords) ──
    header_row_idx = None
    for ri in range(1, min(6, ws.max_row + 1)):
        row_vals = [ws.cell(ri, ci).value for ci in range(1, ws.max_column + 1)]
        combined = ' '.join(str(v).upper() for v in row_vals if v)
        if any(kw in combined for kw in ('SALARY', 'SALARIU', 'SEN', 'PERDIEM', 'PER DIEM')):
            header_row_idx = ri
            break

    if header_row_idx is None:
        raise ValueError("Could not locate a header row containing payroll keywords.")

    data_start = header_row_idx + 1
    headers = [ws.cell(header_row_idx, ci).value for ci in range(1, ws.max_column + 1)]

    # ── 2. Classify each column ──
    name_col = salary_col = perdiem_col = bonus_col = None
    iban_cols: list[int] = []   # 0-based indices, in priority order

    for i, h in enumerate(headers):
        kind = _col_header_type(str(h) if h else '')
        if kind == 'name' and name_col is None:
            name_col = i
        elif kind == 'salary' and salary_col is None:
            salary_col = i
        elif kind == 'perdiem' and perdiem_col is None:
            perdiem_col = i
        elif kind == 'bonus' and bonus_col is None:
            bonus_col = i
        elif kind == 'iban':
            iban_cols.append(i)

    # Defaults if not found by header keyword
    if name_col is None:
        name_col = 0
    if salary_col is None:
        salary_col = 1

    # ── 3. Verify perdiem_col actually contains numbers (not IBANs) ──
    #        In the December format, col C is labelled PERDIEM but holds IBANs.
    if perdiem_col is not None:
        has_numbers = has_ibans = False
        for ri in range(data_start, min(data_start + 8, ws.max_row + 1)):
            v = ws.cell(ri, perdiem_col + 1).value
            if _is_positive_number(v):
                has_numbers = True
            elif _is_iban(v):
                has_ibans = True
        if has_ibans and not has_numbers:
            # It's really an IBAN column despite the header label
            iban_cols.insert(0, perdiem_col)
            perdiem_col = None

    # ── 4. Extract month / year ──
    month, year = _detect_month_year(headers)

    # ── 5. Parse data rows ──
    salary_records: list[dict] = []
    perdiem_records: list[dict] = []
    bonus_records: list[dict] = []
    row_index = 0  # 1-based sequential position among valid salary rows

    for ri in range(data_start, ws.max_row + 1):
        name_val = ws.cell(ri, name_col + 1).value
        if not name_val or str(name_val).strip() == '':
            continue
        name = str(name_val).strip().upper()

        salary_val = ws.cell(ri, salary_col + 1).value
        if not _is_positive_number(salary_val):
            continue

        row_index += 1

        # Salary (pure salary only — bonus is separate)
        salary_amt = round(float(salary_val), 2)

        # Bonus (separate payment, NOT added to salary)
        bonus_amt = None
        if bonus_col is not None:
            bv = ws.cell(ri, bonus_col + 1).value
            if _is_positive_number(bv):
                bonus_amt = round(float(bv), 2)

        # Per diem
        perdiem_amt = None
        if perdiem_col is not None:
            pv = ws.cell(ri, perdiem_col + 1).value
            if _is_positive_number(pv):
                perdiem_amt = round(float(pv), 2)

        # IBAN: try iban_cols in order, then scan all remaining columns
        iban = None
        for ci in iban_cols:
            v = ws.cell(ri, ci + 1).value
            iban = _clean_iban(v)
            if iban:
                break
        if not iban:
            # Scan any column not already assigned
            skip = {name_col, salary_col, perdiem_col, bonus_col} | set(iban_cols)
            for ci in range(ws.max_column):
                if ci in skip:
                    continue
                v = ws.cell(ri, ci + 1).value
                iban = _clean_iban(v)
                if iban:
                    break

        salary_records.append({
            'name': name,
            'amount': salary_amt,
            'iban': iban,
            'index': row_index,
        })

        if perdiem_amt and perdiem_amt > 0:
            perdiem_records.append({
                'name': name,
                'amount': perdiem_amt,
                'iban': iban,
                'salary_index': row_index,
            })

        if bonus_amt and bonus_amt > 0:
            bonus_records.append({
                'name': name,
                'amount': bonus_amt,
                'iban': iban,
                'salary_index': row_index,
            })

    return {
        'month': month,
        'year': year,
        'salary_records': salary_records,
        'perdiem_records': perdiem_records,
        'bonus_records': bonus_records,
    }
